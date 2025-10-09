from datetime import datetime
from io import BytesIO
from typing import List, Optional
import re

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import select
from xlrd import open_workbook

from ..models import Upload, Product, ProductPrice
from ..utils.text import normalize_text
from .pdf_image_importer import import_pdf_or_image


def _infer_columns(headers: List[str]) -> dict:
    lower_cols = {c.strip().lower(): c for c in headers if c is not None}
    name_col = None
    price_col = None
    sku_col: Optional[str] = None
    currency_col: Optional[str] = None

    for key in ["producto", "descripcion", "descripción", "detalle", "articulo", "artículo", "nombre", "name", "producto_nombre", "item"]:
        if key in lower_cols:
            name_col = lower_cols[key]
            break
    if name_col is None and len(headers) >= 1:
        name_col = headers[0]

    for key in ["precio", "price", "unit_price", "mayorista", "wholesale", "valor"]:
        if key in lower_cols:
            price_col = lower_cols[key]
            break
    if price_col is None and len(headers) >= 2:
        price_col = headers[1]

    for key in ["sku", "codigo", "codigo_sku", "code"]:
        if key in lower_cols:
            sku_col = lower_cols[key]
            break

    for key in ["moneda", "currency"]:
        if key in lower_cols:
            currency_col = lower_cols[key]
            break

    return {"name": name_col, "price": price_col, "sku": sku_col, "currency": currency_col}


_num_clean_re = re.compile(r"[^0-9,.-]+")


def try_parse_price(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = _num_clean_re.sub("", s)
    # Heurística: si hay coma y no hay más de un punto, usamos coma como decimal
    if "," in s and s.count(",") == 1 and s.count(".") <= 1:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def choose_price_and_name(headers: List[str], data_rows: List[List[object]]) -> tuple[Optional[str], Optional[str]]:
    # Elegir columna de precio por mayoría de valores numéricos
    header_index = {h: i for i, h in enumerate(headers) if h is not None}
    ncols = len(headers)
    sample = data_rows[:50]
    numeric_counts = [0] * ncols
    text_lengths = [0] * ncols
    alpha_counts = [0] * ncols

    for row in sample:
        for c in range(ncols):
            cell = row[c] if c < len(row) else None
            if try_parse_price(cell) is not None:
                numeric_counts[c] += 1
            s = str(cell).strip() if cell is not None else ""
            if s:
                text_lengths[c] += len(s)
                # Cuenta si hay letras
                if any(ch.isalpha() for ch in s):
                    alpha_counts[c] += 1

    # Price by numeric majority + header hint
    price_candidates = list(range(ncols))
    price_boost = [0] * ncols
    for i, h in enumerate(headers):
        key = (h or "").strip().lower()
        if key in {"precio", "price", "unit_price", "mayorista", "wholesale", "valor"}:
            price_boost[i] += 5
    best_price_idx = max(price_candidates, key=lambda i: (numeric_counts[i] + price_boost[i], numeric_counts[i])) if price_candidates else None

    # Name by header hint then longest text
    name_candidates = list(range(ncols))
    name_boost = [0] * ncols
    for i, h in enumerate(headers):
        key = (h or "").strip().lower()
        if key in {"producto", "descripcion", "nombre", "name", "producto_nombre", "item"}:
            name_boost[i] += 5
    # prefiera columnas con texto alfabético en la mayoría de filas
    best_name_idx = max(
        name_candidates,
        key=lambda i: (
            name_boost[i],
            alpha_counts[i],
            text_lengths[i],
        ),
    ) if name_candidates else None

    price_col = headers[best_price_idx] if best_price_idx is not None else None
    name_col = headers[best_name_idx] if best_name_idx is not None else None
    return price_col, name_col


def find_header_row(rows: List[List[object]]) -> int:
    """Find header row by looking for keyword matches. Requires at least 2 keyword hits."""
    header_keywords = {"producto", "descripcion", "descripción", "nombre", "name", "precio", "price", "sku", "codigo", "código", "moneda", "currency", "presentacion", "presentación"}
    best_idx = -1
    best_score = -1
    max_scan = min(15, len(rows))
    
    for i in range(max_scan):
        row = rows[i]
        non_empty = 0
        keyword_hits = 0
        numeric_like = 0
        
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            non_empty += 1
            sl = s.lower()
            # Exact match or partial match
            if sl in header_keywords:
                keyword_hits += 1
            if try_parse_price(s) is not None:
                numeric_like += 1
        
        # Score formula: prioritize keyword hits, require at least 2 keywords
        # Penalize numeric-heavy rows
        score = keyword_hits * 5 + non_empty - numeric_like * 2
        
        # REQUIRE at least 2 keyword hits to be considered a header
        if non_empty >= 2 and keyword_hits >= 2 and score > best_score:
            best_score = score
            best_idx = i
    
    if best_idx == -1:
        # Fallback: first non-empty row
        for i, row in enumerate(rows[:max_scan]):
            if any((str(c).strip() if c is not None else "") for c in row):
                return i
        return 0
    return best_idx


def extract_provider_name(filename: str) -> str:
    """Extract provider name from upload filename."""
    if not filename:
        return "Proveedor Desconocido"
    
    # Extract first filename if multiple were uploaded together
    fname = filename.split(",")[0].strip()
    # Remove extension
    provider_name = fname.rsplit(".", 1)[0].strip()
    # Remove common patterns like (1), (2), etc
    provider_name = re.sub(r'\(\d+\)$', '', provider_name).strip()
    # Remove extra whitespace
    provider_name = " ".join(provider_name.split())
    
    return provider_name if provider_name else "Proveedor Desconocido"


def _process_product_row(
    name_val: str,
    price_float: float,
    sku_val: Optional[str],
    currency_val: str,
    upload_id: int,
    provider_name: str,
    session: Session,
) -> None:
    """Process a single product row: find or create Product, then create/update ProductPrice."""
    now = datetime.utcnow()
    norm_name = normalize_text(name_val)
    
    # Find or create the Product (by normalized name)
    product = session.execute(
        select(Product).where(Product.normalized_name == norm_name)
    ).scalar_one_or_none()
    
    if product is None:
        # Create new product
        product = Product(
            sku=sku_val if sku_val else None,
            name=name_val,
            normalized_name=norm_name,
            keywords=None,
            created_at=now,
            updated_at=now,
        )
        session.add(product)
        session.flush()  # Get the product.id
    else:
        # Update existing product metadata
        product.updated_at = now
        if sku_val and not product.sku:
            product.sku = sku_val
        session.add(product)
    
    # Find or create ProductPrice for this provider
    existing_price = session.execute(
        select(ProductPrice).where(
            ProductPrice.product_id == product.id,
            ProductPrice.provider_name == provider_name
        )
    ).scalar_one_or_none()
    
    if existing_price:
        # Update existing price
        existing_price.unit_price = round(price_float, 2)
        existing_price.currency = currency_val
        existing_price.source_file_id = upload_id
        existing_price.last_seen_at = now
        existing_price.updated_at = now
        session.add(existing_price)
    else:
        # Create new price entry
        new_price = ProductPrice(
            product_id=product.id,
            source_file_id=upload_id,
            unit_price=round(price_float, 2),
            currency=currency_val,
            provider_name=provider_name,
            last_seen_at=now,
            created_at=now,
            updated_at=now,
        )
        session.add(new_price)


async def import_excels(files: List[UploadFile], session: Session) -> None:
    upload = Upload(filename=", ".join([f.filename for f in files]), uploaded_at=datetime.utcnow())
    session.add(upload)
    session.commit()
    session.refresh(upload)

    total_rows = 0
    total_sheets = 0
    
    # Extract provider name from first file
    provider_name = extract_provider_name(upload.filename)

    for f in files:
        content = await f.read()
        fname = (f.filename or "").lower()

        if fname.endswith((".pdf", ".jpg", ".jpeg", ".png")):
            # Process PDF or image with OCR + GPT-4
            imported = await import_pdf_or_image(
                file_bytes=content,
                filename=f.filename or "unknown",
                upload_id=upload.id,
                provider_name=provider_name,
                session=session,
            )
            total_rows += imported
            # Note: PDFs/images don't have "sheets", so we count as 1 document
            total_sheets += 1
            continue
        elif fname.endswith(".xls"):
            # Parse legacy .xls via xlrd
            book = open_workbook(file_contents=content)
            for sheet in book.sheets():
                total_sheets += 1
                if sheet.nrows == 0:
                    continue
                # Detectar fila de encabezados
                preview = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(min(sheet.nrows, 20))]
                header_row_idx = find_header_row(preview)
                headers = [str(sheet.cell_value(header_row_idx, c)).strip() if sheet.cell_value(header_row_idx, c) is not None else None for c in range(sheet.ncols)]
                mapping = _infer_columns(headers)
                # Mejora: elegir por contenido si falta match de headers
                if mapping["price"] is None or mapping["name"] is None:
                    sample_rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(header_row_idx + 1, min(sheet.nrows, header_row_idx + 61))]
                    price_c, name_c = choose_price_and_name(headers, sample_rows)
                    if mapping["price"] is None:
                        mapping["price"] = price_c
                    if mapping["name"] is None:
                        mapping["name"] = name_c
                name_col = mapping["name"]
                price_col = mapping["price"]
                sku_col = mapping["sku"]
                currency_col = mapping["currency"]
                header_index = {h: i for i, h in enumerate(headers) if h is not None}

                for r in range(header_row_idx + 1, sheet.nrows):
                    try:
                        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                        if name_col is None or price_col is None:
                            continue
                        name_idx = header_index.get(name_col)
                        price_idx = header_index.get(price_col)
                        if name_idx is None or price_idx is None:
                            continue
                        name_cell = row[name_idx] if name_idx < len(row) else None
                        price_cell = row[price_idx] if price_idx < len(row) else None
                        if name_cell is None or str(name_cell).strip() in ("", "nan", "None"):
                            continue
                        if price_cell is None:
                            continue
                        price_float = try_parse_price(price_cell)
                        if price_float is None:
                            continue

                        sku_val = None
                        if sku_col is not None and header_index.get(sku_col) is not None:
                            sku_idx = header_index[sku_col]
                            if sku_idx < len(row) and row[sku_idx] is not None:
                                sku_val = str(row[sku_idx]).strip()

                        currency_val = "ARS"
                        if currency_col is not None and header_index.get(currency_col) is not None:
                            cur_idx = header_index[currency_col]
                            if cur_idx < len(row) and row[cur_idx] is not None:
                                currency_val = str(row[cur_idx]).strip() or "ARS"

                        name_val = str(name_cell).strip()
                        _process_product_row(
                            name_val=name_val,
                            price_float=price_float,
                            sku_val=sku_val,
                            currency_val=currency_val,
                            upload_id=upload.id,
                            provider_name=provider_name,
                            session=session,
                        )
                        total_rows += 1
                    except Exception:
                        continue
                session.commit()
        else:
            # Default to .xlsx via openpyxl
            wb = load_workbook(BytesIO(content), data_only=True)
            for ws in wb.worksheets:
                total_sheets += 1
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_row_idx = find_header_row([list(r) for r in rows[:20]])
                headers = [str(h).strip() if h is not None else None for h in rows[header_row_idx]]
                mapping = _infer_columns(headers)
                if mapping["price"] is None or mapping["name"] is None:
                    sample_rows = [list(r) for r in rows[header_row_idx + 1: min(len(rows), header_row_idx + 61)]]
                    price_c, name_c = choose_price_and_name(headers, sample_rows)
                    if mapping["price"] is None:
                        mapping["price"] = price_c
                    if mapping["name"] is None:
                        mapping["name"] = name_c
                name_col = mapping["name"]
                price_col = mapping["price"]
                sku_col = mapping["sku"]
                currency_col = mapping["currency"]

                header_index = {h: i for i, h in enumerate(headers) if h is not None}

                for row in rows[header_row_idx + 1:]:
                    try:
                        if name_col is None or price_col is None:
                            continue
                        name_idx = header_index.get(name_col)
                        price_idx = header_index.get(price_col)
                        if name_idx is None or price_idx is None:
                            continue
                        name_cell = row[name_idx] if name_idx < len(row) else None
                        price_cell = row[price_idx] if price_idx < len(row) else None
                        if name_cell is None or str(name_cell).strip() in ("", "nan", "None"):
                            continue
                        if price_cell is None:
                            continue
                        price_float = try_parse_price(price_cell)
                        if price_float is None:
                            continue

                        sku_val = None
                        if sku_col is not None and header_index.get(sku_col) is not None:
                            sku_idx = header_index[sku_col]
                            if sku_idx < len(row) and row[sku_idx] is not None:
                                sku_val = str(row[sku_idx]).strip()

                        currency_val = "ARS"
                        if currency_col is not None and header_index.get(currency_col) is not None:
                            cur_idx = header_index[currency_col]
                            if cur_idx < len(row) and row[cur_idx] is not None:
                                currency_val = str(row[cur_idx]).strip() or "ARS"

                        name_val = str(name_cell).strip()
                        _process_product_row(
                            name_val=name_val,
                            price_float=price_float,
                            sku_val=sku_val,
                            currency_val=currency_val,
                            upload_id=upload.id,
                            provider_name=provider_name,
                            session=session,
                        )
                        total_rows += 1
                    except Exception:
                        continue
                session.commit()

    upload.sheet_count = total_sheets
    upload.processed_rows = total_rows
    session.add(upload)
    session.commit()
    print(f"[import] upload_id={upload.id} sheets={total_sheets} rows={total_rows}")


